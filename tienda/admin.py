from django.contrib import admin
from django.contrib.auth.models import User, Group
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Sum
from django.utils import timezone 
from django.utils.html import format_html
from django.urls import path, reverse
from django.shortcuts import redirect

# Importación de todos los modelos del sistema
from .models import Categoria, Producto, MovimientoCaja, Deuda, Abono, Venta, DetalleVenta, VentaCredito

# --- CONFIGURACIÓN DEL SITIO ADMINISTRATIVO PERSONALIZADO ---

class ElZocoAdminSite(admin.AdminSite):
    site_header = "El Zoco Market - Panel de Control"
    index_title = "Administración del Sistema"
    index_template = 'admin/dashboard_zoco.html'

    def index(self, request, extra_context=None):
        from .models import MovimientoCaja, Deuda, Abono, VentaCredito
        from django.db.models import Sum
        from django.utils import timezone
        
        hoy = timezone.now().date()
        
        # --- CÁLCULOS DE CAJA ---
        ingresos_totales = MovimientoCaja.objects.filter(tipo='INGRESO').aggregate(total=Sum('monto'))['total'] or 0
        egresos_totales = MovimientoCaja.objects.filter(tipo='GASTO').aggregate(total=Sum('monto'))['total'] or 0
        
        # --- CÁLCULOS DE DEUDAS (PROVEEDORES) ---
        deudas_qs = Deuda.objects.all()
        total_deuda_original = deudas_qs.aggregate(total=Sum('monto_total'))['total'] or 0
        pendiente_proveedores = sum(d.saldo_pendiente for d in deudas_qs)
        total_pagado_deudas = Abono.objects.filter(deuda__isnull=False, pagado=True).aggregate(total=Sum('monto'))['total'] or 0

        # --- CÁLCULOS DE VENTAS A CRÉDITO (CLIENTES) ---
        creditos_clientes_qs = VentaCredito.objects.all()
        
        # 1. Totales Financieros Coincidentes
        total_credito_original_clientes = creditos_clientes_qs.aggregate(total=Sum('monto_total'))['total'] or 0
        total_por_cobrar_clientes = sum(c.saldo_pendiente for c in creditos_clientes_qs)
        total_cobrado_clientes = Abono.objects.filter(venta_credito__isnull=False, pagado=True).aggregate(total=Sum('monto'))['total'] or 0

        # 2. Conteo de Clientes Activos y Vencidos
        clientes_pendiente_conteo = 0
        clientes_vencidos_conteo = 0

        for credito in creditos_clientes_qs:
            saldo = credito.saldo_pendiente
            if saldo > 0:
                clientes_pendiente_conteo += 1
                
                # Verificamos si tiene plazos programados sin pagar cuya fecha ya expiró
                tiene_plazos_vencidos = credito.abonos.filter(
                    pagado=False,
                    fecha__lt=timezone.now() # La fecha programada es menor a hoy
                ).exists()
                
                if tiene_plazos_vencidos:
                    clientes_vencidos_conteo += 1

        extra_context = extra_context or {}
        extra_context.update({
            'balance': ingresos_totales - egresos_totales,
            'ingresos_totales': ingresos_totales,
            'egresos_totales': egresos_totales,
            
            # Datos de Proveedores
            'pendiente_proveedores': pendiente_proveedores,
            'total_deuda_original': total_deuda_original,
            'total_pagado_deudas': total_pagado_deudas,
            
            # Nuevos Datos de Clientes para el HTML
            'total_por_cobrar_clientes': total_por_cobrar_clientes,
            'total_credito_original_clientes': total_credito_original_clientes,
            'total_cobrado_clientes': total_cobrado_clientes,
            'clientes_pendiente_conteo': clientes_pendiente_conteo,
            'clientes_vencidos_conteo': clientes_vencidos_conteo,
        })
        
        return super().index(request, extra_context)

    def get_app_list(self, request, app_label=None):
        app_dict = self._build_app_dict(request, app_label)
        
        auth_app = app_dict.get('auth')
        tienda_app = app_dict.get('tienda')
        
        
        final_list = []

        # 1. INSERTAR AUTENTICACIÓN PRIMERO
        if auth_app:
            final_list.append(auth_app)

        # 2. CONSTRUIR GRUPOS SEPARADOS (TIENDA Y FINANZAS)
        if tienda_app:
            modelos = tienda_app['models']
            
            # Separar los modelos por su naturaleza conceptual
            modelos_tienda = [m for m in modelos if m['object_name'] in ['Venta', 'Producto', 'Categoria']]
            modelos_finanzas = [m for m in modelos if m['object_name'] in ['Deuda', 'VentaCredito', 'MovimientoCaja']]

            # Agregar Bloque Tienda
            if modelos_tienda:
                tienda_group = tienda_app.copy()
                tienda_group['name'] = 'Tienda'
                tienda_group['models'] = modelos_tienda
                final_list.append(tienda_group)
            
            # Agregar Bloque Finanzas (Incluyendo Deudas y Ventas a Crédito)
            if modelos_finanzas:
                finanzas_group = tienda_app.copy()
                finanzas_group['name'] = 'Finanzas'
                finanzas_group['app_label'] = 'finanzas'
                finanzas_group['models'] = modelos_finanzas
                final_list.append(finanzas_group)

        # 3. Cualquier otra App de terceros
        for app in app_dict.values():
            if app['app_label'] not in ['auth', 'tienda']:
                final_list.append(app)

        return final_list

# Instancia del sitio personalizado
admin_site = ElZocoAdminSite(name='elzoco_admin')


# --- ACCIONES DE EXPORTACIÓN (EXCEL CUSTOM) ---

def generar_excel_creditos_base(queryset, filename, is_proveedor=True):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"
    
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")
    
    persona_label = 'Proveedor' if is_proveedor else 'Cliente'
    headers = [persona_label, 'Monto Original', 'Abonado', 'Fechas de Abonos', 'Resto Pendiente', '% Avance']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        
    for obj in queryset:
        abonos = obj.abonos.filter(pagado=True)
        total_abonado = sum(a.monto for a in abonos)
        resto = obj.monto_total - total_abonado
        avance = (total_abonado / obj.monto_total) if obj.monto_total > 0 else 0
        fechas_texto = ", ".join([a.fecha.strftime('%d/%m/%Y') for a in abonos])
        
        ws.append([obj.persona, obj.monto_total, total_abonado, fechas_texto if fechas_texto else "Sin abonos", resto, avance])
        
        for cell in ws[ws.max_row]:
            cell.border = border
            if cell.column in [2, 3, 5]: 
                cell.number_format = '"$"#,##0'
            if cell.column == 6:
                cell.number_format = '0%'
                cell.alignment = center_align
                if avance >= 1:
                    cell.fill = red_fill
                    cell.font = white_font
                    
    wb.save(response)
    return response

@admin.action(description="💰 Exportar Deudas Seleccionadas (Excel)")
def exportar_deudas_excel_custom(modeladmin, request, queryset):
    return generar_excel_creditos_base(queryset, "Reporte_Deudas_Proveedores_Zoco.xlsx", is_proveedor=True)

@admin.action(description="💰 Exportar Ventas a Crédito Seleccionadas (Excel)")
def exportar_clientes_excel_custom(modeladmin, request, queryset):
    return generar_excel_creditos_base(queryset, "Reporte_Ventas_Credito_Clientes.xlsx", is_proveedor=False)


# --- CONFIGURACIÓN DE INLINES ---

class AbonoInline(admin.TabularInline):
    model = Abono
    extra = 1
    fields = ('monto', 'fecha', 'pagado', 'imprimir_ticket_abono')
    readonly_fields = ('imprimir_ticket_abono',)

    def imprimir_ticket_abono(self, instance):
        if instance.id and instance.pagado:
            url = reverse('ticket_abono', args=[instance.id])
            return format_html(
                '<a class="button" href="{}" target="_blank" style="background-color: #6c757d; color: white; padding: 3px 8px; border-radius: 4px; text-decoration: none;">'
                ' 🖨️ Reimprimir Ticket'
                '</a>', url
            )
        return "No pagado / No guardado"
    imprimir_ticket_abono.short_description = "Acción"


class DetalleVentaInline(admin.TabularInline):
    model = DetalleVenta
    extra = 0
    readonly_fields = ('producto', 'descripcion', 'cantidad', 'precio_unitario', 'subtotal')
    can_delete = False


# --- REGISTROS EN EL SITIO DE ADMINISTRACIÓN PERSONALIZADO ---

@admin.register(Categoria, site=admin_site)
class CategoriaAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'slug')
    prepopulated_fields = {'slug': ('nombre',)}


@admin.register(Producto, site=admin_site)
class ProductoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'precio', 'stock', 'categoria', 'disponible')
    search_fields = ('nombre',)
    list_filter = ('categoria',)


@admin.register(MovimientoCaja, site=admin_site)
class MovimientoCajaAdmin(admin.ModelAdmin):
    list_display = ('tipo', 'descripcion', 'monto', 'fecha', 'imprimir_ticket')
    
    def imprimir_ticket(self, obj):
        return format_html('<a class="button" href="/tienda/ticket/venta/{}/" target="_blank">🎟️ Ticket</a>', obj.id)


# --- MÓDULO FINANCIERO 1: CUENTAS POR PAGAR (PROVEEDORES) ---

@admin.register(Deuda, site=admin_site)
class DeudaAdmin(admin.ModelAdmin):
    actions = [exportar_deudas_excel_custom]
    readonly_fields = ('monto_por_pago',)
    fieldsets = (
        ('Información General', {'fields': ('persona', 'monto_total', 'fecha_inicio')}),
        ('Programación de Pagos', {'fields': ('cantidad_pagos', 'periodicidad_dias', 'monto_por_pago')}),
    )
    # SE REMOVIÓ 'imprimir_ticket_boton' DE AQUÍ PARA QUITAR EL BOTÓN DE LA TABLA INICIAL[cite: 10]
    list_display = ('persona', 'monto_total', 'saldo_pendiente', 'pago_actual', 'monto_por_pago', 'avance_pago')
    search_fields = ('persona', 'monto_total')
    list_filter = ('persona', 'monto_total')
    inlines = [AbonoInline]

    def changelist_view(self, request, extra_context=None):
        deudas_qs = Deuda.objects.all()
        total_original = int(deudas_qs.aggregate(total=Sum('monto_total'))['total'] or 0)
        total_pagado = int(Abono.objects.filter(deuda__isnull=False, pagado=True).aggregate(total=Sum('monto'))['total'] or 0)
        total_pendiente = int(sum(d.saldo_pendiente for d in deudas_qs))

        extra_context = extra_context or {}
        extra_context.update({
            'total_deuda_original': total_original,
            'total_pagado_deudas': total_pagado,
            'pendiente_proveedores': total_pendiente,
        })
        return super().changelist_view(request, extra_context=extra_context)
    
    def pago_actual(self, obj):
        pagados = obj.abonos.filter(pagado=True).count()
        total = obj.cantidad_pagos
        if total > 1:
            if pagados >= total: 
                return format_html('<span style="color: #15ff00; font-weight: bold;">Completado</span>')
            return format_html("Pago <b>{}</b> de <b>{}</b>", pagados, total)
        return "Pago Único"
    pago_actual.short_description = "Esquema"

    def avance_pago(self, obj):
        total = float(obj.monto_total)
        pendiente = float(obj.saldo_pendiente)
        pagado = total - pendiente
        porcentaje = int((pagado / total) * 100) if total > 0 else 0
        color = "#ff4d4d" if porcentaje < 30 else "#f57e06" if porcentaje < 70 else "#19850f"
        return format_html('<div style="background-color: {}; color: white; width: 40px; height: 40px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-weight: bold;">{}%</div>', color, porcentaje)


# --- MÓDULO FINANCIERO 2: VENTAS A CRÉDITO (CLIENTES) ---

@admin.register(VentaCredito, site=admin_site)
class VentaCreditoAdmin(admin.ModelAdmin):
    actions = [exportar_clientes_excel_custom]
    readonly_fields = ('monto_por_pago', 'venta')
    fieldsets = (
        ('Información del Cliente', {'fields': ('persona', 'monto_total', 'fecha_inicio', 'venta')}),
        ('Plazos y Financiamiento', {'fields': ('cantidad_pagos', 'periodicidad_dias', 'monto_por_pago')}),
    )
    # SE REMOVIÓ 'imprimir_ticket_boton' DE AQUÍ TAMBIÉN PARA QUITAR EL BOTÓN DE LA TABLA INICIAL[cite: 10]
    list_display = ('persona', 'monto_total', 'saldo_pendiente', 'pago_actual', 'monto_por_pago', 'avance_pago', 'ver_origen_pos')
    search_fields = ('persona', 'monto_total')
    list_filter = ('persona',)
    inlines = [AbonoInline]

    def changelist_view(self, request, extra_context=None):
        creditos_qs = VentaCredito.objects.all()
        total_original = int(creditos_qs.aggregate(total=Sum('monto_total'))['total'] or 0)
        total_cobrado = int(Abono.objects.filter(venta_credito__isnull=False, pagado=True).aggregate(total=Sum('monto'))['total'] or 0)
        total_pendiente = int(sum(c.saldo_pendiente for c in creditos_qs))

        extra_context = extra_context or {}
        extra_context.update({
            'total_deuda_original': total_original,
            'total_pagado_deudas': total_cobrado,
            'pendiente_proveedores': total_pendiente,
        })
        return super().changelist_view(request, extra_context=extra_context)

    def pago_actual(self, obj):
        pagados = obj.abonos.filter(pagado=True).count()
        total = obj.cantidad_pagos
        if total > 1:
            if pagados >= total: 
                return format_html('<span style="color: #15ff00; font-weight: bold;">Completado</span>')
            return format_html("Pago <b>{}</b> de <b>{}</b>", pagados, total)
        return "Pago Único"
    pago_actual.short_description = "Esquema"

    def avance_pago(self, obj):
        total = float(obj.monto_total)
        pendiente = float(obj.saldo_pendiente)
        pagado = total - pendiente
        porcentaje = int((pagado / total) * 100) if total > 0 else 0
        color = "#ff4d4d" if porcentaje < 30 else "#f57e06" if porcentaje < 70 else "#19850f"
        return format_html('<div style="background-color: {}; color: white; width: 40px; height: 40px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-weight: bold;">{}%</div>', color, porcentaje)

    def ver_origen_pos(self, obj):
        if obj.venta:
            url = reverse('admin:tienda_venta_change', args=[obj.venta.id])
            return format_html('<a href="{}">🛒 Venta #{}</a>', url, obj.venta.folio)
        return "Carga Manual"
    ver_origen_pos.short_description = "Origen POS"


# --- HISTORIAL GENERAL DE VENTAS (Cajas/POS) ---

@admin.register(Venta, site=admin_site)
class VentaAdmin(admin.ModelAdmin):
    list_display = ('folio', 'vendedor', 'total_destacado', 'forma_pago', 'fecha', 'reimprimir_ticket')
    readonly_fields = ('folio', 'fecha', 'total', 'forma_pago', 'cliente', 'vendedor')
    inlines = [DetalleVentaInline]
    list_filter = ('forma_pago', 'fecha')
    search_fields = ('folio', 'cliente')
    
    def reimprimir_ticket(self, obj):
        url = reverse('ticket_venta', args=[obj.id])
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background-color: #444; color: white; padding: 5px 10px; border-radius: 4px; text-decoration: none;">'\
            ' <i class="fas fa-print"></i> Reimprimir'\
            '</a>', url
        )
    reimprimir_ticket.short_description = "Acción"

    def total_destacado(self, obj):
        total_formateado = "{:,.0f}".format(obj.total)
        return format_html('<span style="font-size: 18px; font-weight: bold;">${}</span>', total_formateado)
    total_destacado.short_description = "Total"

# Registrar Autenticación para mantener el orden estricto de get_app_list
admin_site.register(User)
admin_site.register(Group)